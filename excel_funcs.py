import os
import time
import openpyxl
from openpyxl.styles import Font, Style, colors, PatternFill



def get_creds(selected_region):
	""" Retrieves Username and Password from the Excel sheet if present, else goes silent """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Get the sheet - Connections
	sheet = wb.get_sheet_by_name("Connections")

	# List of credentials
	creds = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	# Get the selected region's Username and Password
	for i in range(2, max_rows + 1):
		if selected_region == sheet.cell(row=i, column=1).value:
			creds.append(sheet.cell(row=i, column=5).value)
			creds.append(sheet.cell(row=i, column=6).value)
		break

	# Remove None from the List
	final_creds = [i for i in creds if i is not None]

	# Return the final list
	return final_creds



def get_regions():
	""" Returns the regions """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Get the sheet - Connections
	sheet = wb.get_sheet_by_name("Connections")

	# List of regions
	regions_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Looping in through the rows and adding the regions to the list
	for i in range(2, max_rows + 1):
		regions_list.append(sheet.cell(row=i, column=1).value)

	# Returns the list of regions
	return regions_list

def get_sub_regions(selected_flag):
	""" Returns a list of sub regions """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Get the sheet - Connections
	sheet = wb.get_sheet_by_name("Connections")

	# List of regions
	regions_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Looping in through the rows and adding the regions to the list
	for i in range(2, max_rows + 1):
		if selected_flag[:-1] == sheet.cell(row=i, column=7).value:
			regions_list.append(sheet.cell(row=i, column=1).value)

	# Returns the list of regions
	return regions_list


def get_db(selected_region):
	""" Returns all the connection details for a specific region given in the function argument """

	# Variable declarations
	k = 1

	# Workbook Object 
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Sheet object from the Workbook
	sheet = wb.get_sheet_by_name('Connections')

	# Max filled rows
	max_rows = sheet.max_row

	# Max filled columns
	max_cols = sheet.max_column

	# Traversing through each row and comparing the region
	for i in range(2, max_rows + 1):
		if selected_region == sheet.cell(row=i, column=1).value:
			k = 0
			db = sheet.cell(row=i, column=4).value
		else:
			pass

		if k == 0:
			break

	# return the database name
	return db


def get_config(selected_region):
	""" Returns all the connection details for a specific region given in the function argument """

	# Config list
	config_list = []

	# Variable declarations
	k = 1

	# Workbook Object 
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Sheet object from the Workbook
	sheet = wb.get_sheet_by_name('Connections')

	# Max filled rows
	max_rows = sheet.max_row

	# Max filled columns
	max_cols = sheet.max_column

	# Traversing through each row and comparing the region
	for i in range(2, max_rows + 1):
		if selected_region == sheet.cell(row=i, column=1).value:
			k = 0
			for j in range(1, max_cols + 1):
				config_list.append(sheet.cell(row=i, column=j).value)
		else:
			pass

		if k == 0:
			break

	final_config_list = [i for i in config_list if i is not None]
	# return the database name
	return final_config_list

def fetch_table_names(selected_region):
	""" Fetches the names of the Tables from the Excel Sheet according to the region selected """

	# List declared for collecting table names
	selected_region_tables = []

	# Workbook object created
	wb = openpyxl.load_workbook("DB_Check_Configure.xlsx")

	# sheet object created
	sheet = wb.get_sheet_by_name("Tables")

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	# Traversing each cell to fetch the table names
	for i in range(2, max_rows + 1):
		if selected_region == sheet.cell(row=i, column=1).value:
			selected_region_tables.append(sheet.cell(row=i, column=2).value)
		else:
			pass

	# Returning a list of table names
	return selected_region_tables

def get_input_params(selected_region, selected_table):
	""" Returns the list of input parameters for the Select Query """

	# List to collect the input parameters from the Excel file
	params_list = []

	# WorkBook object
	wb = openpyxl.load_workbook("DB_Check_Configure.xlsx")

	# Sheet object
	sheet = wb.get_sheet_by_name("Tables")

	# Maximum filled rows
	max_rows = sheet.max_row

	# Maximum filled columns
	max_cols = sheet.max_column

	# Iterating through each row and appending the input parameters
	for i in range(2, max_rows + 1):
		if (selected_region == sheet.cell(row=i, column=1).value and selected_table == sheet.cell(row=i, column=2).value):
			for j in range(4, max_cols + 1):
				params_list.append(sheet.cell(row=i, column=j).value)
		else:
			pass

	# Removing None from the list, and preparing a final list
	final_params_list = [i for i in params_list if i is not None]

	# Returning a correct list of input parameters
	return final_params_list

def get_validations(selected_table):
	""" Returns the Column name to be validated and its Expected Value """

	# Empty dictionary variable
	dict_resp = {}

	flag = 1

	# Table name
	table_name = selected_table.split(".")[1]

	# Workbook Object
	wb = openpyxl.load_workbook("DB_Check_Configure.xlsx")

	# Sheet Object
	sheet = wb.get_sheet_by_name("Validation")

	# Maximum filled rows
	max_rows = sheet.max_row

	# Maximum filled columns
	max_cols = sheet.max_column

	# Iterating through each row
	for i in range(2, max_rows + 1):
		if table_name == sheet.cell(row=i, column=1).value:
			flag = 0

			# Traversing the complete row
			for j in range(2, max_cols + 1, 2):

				# Continue the loop if the cell is empty
				if sheet.cell(row=i, column=j).value == None:
					continue
				else:

					# Else add value to the dictionary
					vals_list = (sheet.cell(row=i, column=j + 1).value).split(",")
					dict_resp[sheet.cell(row=i, column=j).value] = vals_list

		if flag == 0:
			break

	return dict_resp

def get_usercolumnsel(selected_region, input_type, reg_region_list = None):
	""" Returns user selection column as a dict """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Get the sheet - Tables
	sheet = wb.get_sheet_by_name("Tables")

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Dummy dict
	mega_list = []

	all_regions = []
	
	# Appending the current regin to the list of regions
	if reg_region_list != None:
		all_regions = [i for i in reg_region_list]
	all_regions.append(selected_region)
	

	# Looping in through the rows and adding the regions to the list
	for region in all_regions:

		for i in range(2, max_rows + 1):

			if (str(sheet.cell(row=i, column=1).value)).strip() == str(region).strip():
				
				if sheet.cell(row=i, column=2).value != None and sheet.cell(row=i, column=3).value != None:
					temp = []
					temp.append(region)
					temp.append(sheet.cell(row=i, column=2).value)

					all_inputs = (sheet.cell(row=i, column=3).value).split(";")

					for i in range(0, len(all_inputs)):

						# If the input type is selected as FileID or FileName etc
						if str(input_type).strip().upper() == str(all_inputs[i].split(":")[0]).strip().upper():
							temp.append(str(all_inputs[i].split(":")[1]).strip().upper())
							mega_list.append(temp)
							

	# Returns the list of regions
	return mega_list

def retrieve_inputs_from_file(col_name):
	"""  Retrieves the values from the user input sheet """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('Input\\User_Input_Sheet.xlsx')

	# Get the sheet - Tables
	sheet = wb.get_sheet_by_name("Inputs")

	my_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	for i in range(2, max_rows + 1):
		for j in range(1, max_cols + 1):
			if str(sheet.cell(row=1, column=j).value).strip() == str(col_name).strip():
				my_list.append(str(sheet.cell(row=i, column=j).value).strip())

	return my_list

def get_user_input_types():
	""" Returns Column H of the Connections sheet """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('DB_Check_Configure.xlsx')

	# Get the sheet - Tables
	sheet = wb.get_sheet_by_name("Connections")

	user_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	for i in range(2, max_rows + 1):
		if sheet.cell(row=i, column=8).value != None:
			user_list.append(str(sheet.cell(row=i, column=8).value).strip())


	return user_list

def create_result_file(str_path, check_path, sheet_title):
	""" Creates regression result files """

	if not os.path.exists(str(str_path)):
		os.makedirs(str(str_path))

	if not os.path.exists(str(check_path)):

		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = str(sheet_title)

		wb.save(str(check_path))
		# wb.close()

		return 0

def write_result(file_path, sheet_title, search_param = None, table_name= None, expceted_value= None, actual_value= None, status=None):
	""" Writes the result in a given file and sheet """
	
	# Work Book from current work directory
	wb = openpyxl.load_workbook(file_path)

	# Get the sheet - Tables
	sheet = wb.get_sheet_by_name(sheet_title)

	max_rows = sheet.max_row

	yfill = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
	gft = Font(color='009999')
	rft = Font(color=colors.RED)

	if max_rows == 1:
		sheet.cell(row=max_rows, column=1).value = "S. No."
		sheet.cell(row=max_rows, column=1).fill = yfill
		sheet.cell(row=max_rows, column=2).value = "Search Parameter"
		sheet.cell(row=max_rows, column=2).fill = yfill
		sheet.cell(row=max_rows, column=3).value = "Table Name"
		sheet.cell(row=max_rows, column=3).fill = yfill
		sheet.cell(row=max_rows, column=4).value = "Expected Value"
		sheet.cell(row=max_rows, column=4).fill = yfill
		sheet.cell(row=max_rows, column=5).value = "Actual Value"
		sheet.cell(row=max_rows, column=5).fill = yfill

		sheet.cell(row=max_rows + 1, column=1).value = max_rows
	else:
		sheet.cell(row=max_rows, column=2).value = str(search_param).strip()
		sheet.cell(row=max_rows, column=3).value = str(table_name).strip()
		sheet.cell(row=max_rows, column=4).value = str(expceted_value).strip()

		if str(status) == "Pass":
			sheet.cell(row=max_rows, column=5).value = str(actual_value).strip()
			sheet.cell(row=max_rows, column=5).font = gft
		elif str(status) == "Fail":
			sheet.cell(row=max_rows, column=5).value = str(actual_value).strip()
			sheet.cell(row=max_rows, column=5).font = rft

		sheet.cell(row=max_rows + 1, column=1).value = max_rows

	wb.save(str(file_path))
	# wb.close()

	return 0

def create_log_file(str_path, check_path):
	""" Creates and writes logs """

	if not os.path.exists(str(str_path)):
		os.makedirs(str(str_path))

	if not os.path.exists(str(check_path)):

		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = str(sheet_title)

		wb.save(str(check_path))
		# wb.close()

		return 0

# l = create_result_file("C:\Users\msiddiq1\Documents\Test-Python\Project DB\New folder\Result\Ejaz", "Ejaz", "Result", "xlsx")
# print l

# l = write_result("C:\Users\msiddiq1\Documents\Test-Python\Project DB\New folder\Result\\2016.03.11\\2016.03.11 - 17.45.xlsx", "Result")
# print l

# l = os.getcwd()
# l = "%s\\%s" % (str(l), "Input", )
# l = l.replace("\\", "\\\\")
# print l